import re
import time
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
import pdfplumber
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

st.set_page_config(page_title="GCTS - Cruce Objetivos SAFA/SACA/SANA", layout="wide")

APP_TITLE = "Cruce automático: Lista de tráfico (NOP) + Objetivos SAFA/SACA/SANA"
APP_CAPTION = (
    "Sube el PDF NOP y el Excel maestro. La matrícula del Excel maestro tiene prioridad "
    "para determinar el código operativo de cruce; hexdb.io y el código de vuelo se usan "
    "como contraste."
)

OUTPUT_COLUMNS = [
    "Hora", "ARCID", "Aeronave", "Matricula", "ADEP", "ADES", "prefix3",
    "Código externo", "Operador (maestro)", "Tipo objetivo",
    "Inspecciones realizadas", "Objetivo 2026", "Restantes", "Última inspección",
]
INTEGER_COLUMNS = {"Inspecciones realizadas", "Objetivo 2026", "Restantes"}

ATYP_CODES_RAW = """
A124 A140 A148 A158 A19N A20N A21N A225 A306 A30B A310 A318 A319 A320 A321 A332 A333
A337 A338 A339 A342 A343 A345 A346 A359 A35K A388 A3ST A400 A748
AC90 AJ27 AN12 AN24 AN26 AN28 AN30 AN32 AN72 AT43 AT44 AT45 AT46 AT72 AT73 AT75 AT76 ATP
B190 B37M B38M B39M B3XM B461 B462 B463 B703 B712 B720 B721 B722 B732 B733 B734 B735
B736 B737 B738 B739 B741 B742 B743 B744 B748 B74R B74S B752 B753 B762 B763 B764 B772
B773 B778 B779 B77L B77W B788 B789 B78X BA11 BA46 BCS1 BCS3 BE20 BE30 BE33 BE35 BE36
BE40 BE58 BE76 BE95 BE99 BELF BER2 BLCF
C130 C17 C172 C177 C182 C206 C208 C210 C212 C25A C25B C25C C25M C310 C340 C408 C414
C421 C500 C510 C525 C550 C560 C56X C5M C650 C680 C68A C700 C750 C919 CL2T CL30 CL35
CL60 CN35 CRJ1 CRJ2 CRJ7 CRJ9 CRJX CVLT
D228 D328 DA20 DA40 DA42 DA62 DC10 DC85 DC86 DC87 DC91 DC92 DC93 DC94 DC95 DH8A DH8B
DH8C DH8D DHC5 DHC6 DHC7 DHT
E110 E120 E135 E145 E170 E175 E190 E195 E290 E295 E35L E50P E545 E550 E55P E75L E75S EA50
F100 F27 F28 F2TH F406 F50 F70 F900 FA50 FA6X FA7X FA8X
G150 G159 G200 G280 G650 G73T GA5C GA6C GA7C GA8 GALX GL5T GL7T GLEX GLF4 GLF5 GLF6
H25B H25C HDJT I114 IL18 IL62 IL76 IL86 IL96 J328 JS31 JS32 JS41 K35R
L101 L188 L410 LJ31 LJ35 LJ40 LJ45 LJ60 LJ70 LJ75
M20P M20T MD11 MD81 MD82 MD83 MD87 MD88 MD90 MU2 N262 NOMA
P180 P206 P208 P210 P28A P28B P28R P28T P68C P8 P92 PA28 PA34 PA44 PA46 PAY2 PC12 PC24 PRM1
RJ1H RJ70 RJ85 S601 SB20 SC7 SF34 SH33 SH36 SR20 SR22 SU95 SW4
T134 T154 T204 TBM7 TBM8 TBM9 TU34 TU54 WW24 Y12 YK40 YK42 YS11
""".split()
ATYP_CODES = sorted(set(ATYP_CODES_RAW), key=len, reverse=True)
ATYP_PAT = re.compile(r"(" + "|".join(re.escape(c) for c in ATYP_CODES) + r")")

REG_PREFIXES_ALL = """
A9C 4YB 9XR 9A 9G 9H 9J 9K 9L 9M 9N 9Q 9U 9V 9Y 4K 4L 4O 4R 4X
5A 5B 5H 5N 5R 5T 5U 5V 5W 5X 5Y 6O 6V 6W 6Y 7O 7P 7Q 7T 8P 8Q 8R
A2 A3 A4 A5 A6 A7 A8 AP C2 C5 C6 C9 CC CN CP CR CS CU D2 D4 D6
E7 EC EI EJ EK EL EP ES ET EW EX EY EZ H4 HA HB HC HH HI HK HL HP HR HS HZ
JA JU JY LN LQ LR LV LX LY LZ MI MT OB OD OE OH OK OM OO OY PH PJ PK PP PR PT PU PZ
RA RP S2 S5 S7 S9 SE SP ST SU SX T2 T3 T7 T8 T9 TC TF TG TI TJ TL TN TR TS TT TU TY TZ
UK UN UP UR V2 V3 V4 V5 V6 V7 V8 VH VN VP VQ VR VT XA XB XC XT XU XY YA YI YJ YK YL YR YU YV
Z3 ZA ZJ ZK ZM ZP ZQ ZS B C D F G I J M N P T V Z
""".split()
REG_PREFIXES_SORTED = sorted(set(REG_PREFIXES_ALL), key=len, reverse=True)

TTV_PAT = re.compile(r"[A-Za-z]\s?\d{3}\s?\d{2}-")
TIME_PAT_F1 = re.compile(r"^(\d{2}:\d{2})A\s*(.*)$")
FLIGHT_LINE_PAT_F1 = re.compile(r"^\d{2}:\d{2}A")
FLIGHT_START_PAT_F2 = re.compile(r"(\d{2}:\d{2})([AEC])(?=\s?(?:[A-Z]{1,4}\s?)?[A-Z]{1,4}\d)")
EXPECTED_TOTAL_PAT = re.compile(r"-\s*(\d+)\s*Flights", re.IGNORECASE)
AIRPORT_PAT = re.compile(r"^[A-Z0-9]{4}$")

HEXDB_REG_HEX_URL = "https://hexdb.io/reg-hex"
HEXDB_AIRCRAFT_URL = "https://hexdb.io/api/v1/aircraft/{hex}"
HEXDB_REQUEST_TIMEOUT = 5
HEXDB_REQUEST_DELAY = 0.25


@dataclass
class CrossResult:
    tipo: str
    operador: str
    inspecciones: Optional[object]
    objetivo: Optional[object]
    restantes: Optional[object]
    ultima: str
    codigo_externo: str
    discrepancia: bool
    codigo_usado: str


def _norm_header(value) -> str:
    s = str(value or "")
    s = unicodedata.normalize("NFKD", s)
    return re.sub(r"\s+", " ", "".join(c for c in s if not unicodedata.combining(c))).strip().upper()


def _find_col(headers_norm: Dict[str, int], candidates: List[str]) -> Optional[int]:
    for cand in candidates:
        cand_norm = _norm_header(cand)
        for h_norm, idx in headers_norm.items():
            if cand_norm in h_norm:
                return idx
    return None


def normalize_registration(value) -> str:
    return str(value or "").strip().upper().replace(" ", "").replace("-", "")


def hyphenate_registration(reg_raw: str) -> str:
    reg_raw = (reg_raw or "").strip()
    if not reg_raw:
        return ""
    if re.match(r"^N\d", reg_raw):
        return reg_raw
    for prefix in REG_PREFIXES_SORTED:
        if reg_raw.startswith(prefix) and len(reg_raw) > len(prefix):
            return f"{prefix}-{reg_raw[len(prefix):]}"
    return reg_raw


def to_int_or_none(value):
    try:
        if value is None or (isinstance(value, str) and not value.strip()):
            return None
        f = float(value)
        return None if pd.isna(f) else int(round(f))
    except (TypeError, ValueError):
        return None


def fmt_date(v) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return str(v)


def looks_like_format2(text: str) -> bool:
    for line in text.splitlines():
        line = line.strip()
        if re.match(r"^\d{2}:\d{2}[AEC]", line):
            return True
        if re.match(r"^\d{2}:\d{2}A\s", line):
            return False
    return False


def extract_reg_airports(block: str) -> Tuple[str, str, str]:
    compact = block.replace(" ", "").replace(">", "")
    if len(compact) >= 8:
        return compact[:-8], compact[-8:-4], compact[-4:]
    if len(compact) > 4:
        return "", compact[:4], compact[4:]
    return "", compact, ""


def parse_one_flight_chunk(hora: str, rest: str) -> Tuple[Optional[Dict[str, str]], Optional[str]]:
    rest = rest.lstrip()
    atyp_match = ATYP_PAT.search(rest)
    if not atyp_match:
        return None, "Tipo de aeronave no reconocido"
    prefix = rest[:atyp_match.start()]
    atyp = atyp_match.group(1)
    remainder = rest[atyp_match.end():]
    dm = re.search(r"\d", prefix)
    if dm is None:
        return None, "Indicativo sin número de vuelo (posible aeronave privada/GA)"
    airline_code = prefix[:dm.start()][-3:]
    arcid = (airline_code + prefix[dm.start():]).strip()
    anchor = TTV_PAT.search(remainder)
    reg_block = remainder[:anchor.start()] if anchor else remainder
    reg_raw, adep, ades = extract_reg_airports(reg_block)
    if not AIRPORT_PAT.match(adep or "") or not AIRPORT_PAT.match(ades or ""):
        return None, "Formato de ADEP/ADES inesperado tras el parsing"
    return {
        "Hora": hora, "ARCID": arcid, "Aeronave": atyp,
        "Matricula": hyphenate_registration(reg_raw), "ADEP": adep,
        "ADES": ades, "prefix3": airline_code,
    }, None


def parse_format2(raw_lines: List[str]) -> Tuple[pd.DataFrame, List[Dict[str, str]]]:
    rows, unparsed = [], []
    for line in raw_lines:
        matches = list(FLIGHT_START_PAT_F2.finditer(line))
        for i, match in enumerate(matches):
            end = matches[i + 1].start() if i + 1 < len(matches) else len(line)
            parsed, reason = parse_one_flight_chunk(match.group(1), line[match.end():end])
            if parsed:
                rows.append(parsed)
            else:
                unparsed.append({"Hora": match.group(1), "raw": line[match.end():end].strip()[:80], "motivo": reason})
    return pd.DataFrame(rows), unparsed


def parse_format1(raw_lines: List[str]) -> Tuple[pd.DataFrame, List[Dict[str, str]]]:
    merged, current = [], None
    for line in raw_lines:
        if FLIGHT_LINE_PAT_F1.match(line):
            if current:
                merged.append(current)
            current = line
        elif current:
            current += " " + line
    if current:
        merged.append(current)
    rows, unparsed = [], []
    for line in merged:
        m = TIME_PAT_F1.match(line)
        if not m:
            continue
        hora, rest = m.group(1), m.group(2)
        atyp_match = ATYP_PAT.search(rest)
        if not atyp_match:
            unparsed.append({"Hora": hora, "raw": rest[:80], "motivo": "Tipo de aeronave no reconocido"})
            continue
        prefix = rest[:atyp_match.start()].strip()
        if not re.search(r"\d", prefix):
            unparsed.append({"Hora": hora, "raw": rest[:80], "motivo": "Indicativo sin número de vuelo"})
            continue
        rem = rest[atyp_match.end():].replace(" ", "")
        rows.append({"Hora": hora, "ARCID": prefix, "Aeronave": atyp_match.group(1),
                     "Matricula": hyphenate_registration(rem[:5]), "ADEP": rem[5:9],
                     "ADES": rem[9:13], "prefix3": prefix[:3]})
    return pd.DataFrame(rows), unparsed


def parse_pdf_flights(pdf_bytes: bytes) -> Tuple[pd.DataFrame, pd.DataFrame]:
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        raw_lines = []
        for page in pdf.pages:
            raw_lines.extend([line.strip() for line in (page.extract_text() or "").splitlines() if line.strip()])
    full_text = "\n".join(raw_lines)
    flights, undetected = parse_format2(raw_lines) if looks_like_format2(full_text) else parse_format1(raw_lines)
    return flights, pd.DataFrame(undetected, columns=["Hora", "raw", "motivo"])


def extract_expected_total(pdf_bytes: bytes) -> Optional[int]:
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            m = EXPECTED_TOTAL_PAT.search(page.extract_text() or "")
            if m:
                return int(m.group(1))
    return None


@st.cache_data(show_spinner=False, ttl=60 * 60 * 24)
def hexdb_lookup_operator_code(registration: str) -> Optional[str]:
    """Optional external reference only. It never overrides a matching
    registration from the Excel master, which is the authoritative source
    for this inspection-planning workflow."""
    reg = (registration or "").strip().upper()
    if not reg:
        return None
    try:
        r = requests.get(HEXDB_REG_HEX_URL, params={"reg": reg}, timeout=HEXDB_REQUEST_TIMEOUT)
        time.sleep(HEXDB_REQUEST_DELAY)
        if r.status_code != 200:
            return None
        hex_code = r.text.strip()
        if not hex_code or "not found" in hex_code.lower() or len(hex_code) > 8:
            return None
        r2 = requests.get(HEXDB_AIRCRAFT_URL.format(hex=hex_code), timeout=HEXDB_REQUEST_TIMEOUT)
        time.sleep(HEXDB_REQUEST_DELAY)
        if r2.status_code != 200:
            return None
        return (str(r2.json().get("OperatorFlagCode") or "").strip().upper() or None)
    except Exception:
        return None


def build_external_operator_map(registrations: List[str]) -> Dict[str, Optional[str]]:
    regs = sorted({str(x).strip().upper() for x in registrations if str(x).strip()})
    output = {}
    if not regs:
        return output
    progress = st.progress(0.0, text=f"Consultando hexdb.io: 0 / {len(regs)} matrículas")
    for i, reg in enumerate(regs, start=1):
        output[reg] = hexdb_lookup_operator_code(reg)
        if i % 5 == 0 or i == len(regs):
            progress.progress(i / len(regs), text=f"Consultando hexdb.io: {i} / {len(regs)} matrículas")
    progress.empty()
    return output


def load_sheet(xlsx_bytes: bytes, name: str, max_col: int = 60):
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb[name]
    rows = list(ws.iter_rows(min_col=1, max_col=max_col, values_only=True))
    wb.close()
    return rows[0] if rows else (), rows[1:]


def find_sheet_name(names: List[str], keywords: List[str]) -> Optional[str]:
    for name in names:
        normalized = _norm_header(name)
        if all(_norm_header(k) in normalized for k in keywords):
            return name
    return None


def tokens_from_registration_cell(value) -> List[str]:
    """Extracts registration tokens from cells that contain registrations
    optionally concatenated with type codes, separated by comma/whitespace.
    For example 'EC-NZRB738, EC-OBKB738' yields ECNZR and ECOBK."""
    text = str(value or "").upper()
    candidates = re.findall(r"(?:[A-Z0-9]{1,3}-[A-Z0-9]{2,6}|N\d[A-Z0-9]{1,5})", text)
    return [normalize_registration(x) for x in candidates if normalize_registration(x)]


def build_master_maps(xlsx_bytes: bytes):
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    sheetnames = wb.sheetnames
    wb.close()
    l1_sheet = find_sheet_name(sheetnames, ["LAYER 1"])
    l2_sheet = find_sheet_name(sheetnames, ["LAYER 2"])
    sana_sheet = find_sheet_name(sheetnames, ["SANA"])
    matriculas_sheet = find_sheet_name(sheetnames, ["MATR"]) or find_sheet_name(sheetnames, ["MATRICULAS"])
    icao_sheet = find_sheet_name(sheetnames, ["ICAO"]) or find_sheet_name(sheetnames, ["CODE"])

    icao_map, l1_map, l2_map, sana_map, matriculas_map, registration_code_map = {}, {}, {}, {}, {}, {}

    if icao_sheet:
        headers, rows = load_sheet(xlsx_bytes, icao_sheet, 2)
        for row in rows:
            if len(row) > 1 and row[0] and row[1]:
                icao_map[str(row[1]).strip().upper()] = str(row[0]).strip()

    def read_objective_sheet(sheet, target_map):
        if not sheet:
            return
        headers, rows = load_sheet(xlsx_bytes, sheet)
        norm = {_norm_header(v): i for i, v in enumerate(headers) if v}
        code_i = _find_col(norm, ["3LC", "OACI"])
        op_i = _find_col(norm, ["OPERATOR NAME", "OPERATOR L2", "OPERADOR"])
        done_i = _find_col(norm, ["PROGRESS", "DONE", "INSPECCIONES REALIZ"])
        obj_i = _find_col(norm, ["MEAN TARGET", "OBJECTIVE 2026", "INSPECCIONES OBJETIVO"])
        rem_i = _find_col(norm, ["REMAINING", "FALTANTES"])
        last_i = _find_col(norm, ["LAST INSPECTION SPAIN", "LAST INSPECTION", "ULTIMA INSPECCION"])
        reg_i = _find_col(norm, ["REGISTRATION MARKINGS", "REG. MARKINGS", "MATRICULAS"])
        for row in rows:
            code = row[code_i] if code_i is not None and code_i < len(row) else None
            if not code:
                continue
            code = str(code).strip().upper()
            item = {
                "operator": row[op_i] if op_i is not None and op_i < len(row) else "",
                "done": row[done_i] if done_i is not None and done_i < len(row) else None,
                "objective": row[obj_i] if obj_i is not None and obj_i < len(row) else None,
                "remaining": row[rem_i] if rem_i is not None and rem_i < len(row) else None,
                "last": row[last_i] if last_i is not None and last_i < len(row) else None,
            }
            target_map[code] = item
            if reg_i is not None and reg_i < len(row):
                for reg in tokens_from_registration_cell(row[reg_i]):
                    registration_code_map.setdefault(reg, code)

    read_objective_sheet(l1_sheet, l1_map)
    read_objective_sheet(l2_sheet, l2_map)
    read_objective_sheet(sana_sheet, sana_map)

    if matriculas_sheet:
        headers, rows = load_sheet(xlsx_bytes, matriculas_sheet, 25)
        norm = {_norm_header(v): i for i, v in enumerate(headers) if v}
        reg_i = _find_col(norm, ["MATRICULA SIN", "MATRICULA", "REGISTRATION"])
        op_i = _find_col(norm, ["OPERADOR", "OPERATOR"])
        for row in rows:
            reg = row[reg_i] if reg_i is not None and reg_i < len(row) else None
            op = row[op_i] if op_i is not None and op_i < len(row) else None
            if reg and op:
                matriculas_map[normalize_registration(reg)] = str(op).strip()

    return icao_map, l1_map, l2_map, sana_map, matriculas_map, registration_code_map


def resolve_operator_code(registration: str, flight_code: str, external_code: Optional[str], registration_code_map: Dict[str, str]):
    """Priority: Excel registration mapping > external API > flight code.

    The Excel mapping is authoritative. Thus EC-NZR maps to OVA when OVA's
    'Registration Markings' includes EC-NZR, even if AEA appears in the
    flight plan and is returned by an external data provider.
    """
    reg_key = normalize_registration(registration)
    flight = str(flight_code or "").strip().upper()
    external = str(external_code or "").strip().upper()
    excel_code = registration_code_map.get(reg_key, "")
    if excel_code:
        discrepancy = bool((flight and flight != excel_code) or (external and external != excel_code))
        return excel_code, discrepancy
    if external:
        return external, bool(flight and flight != external)
    return flight, False


def cross_reference(row: pd.Series, maps, external_codes: Dict[str, Optional[str]]) -> CrossResult:
    icao_map, l1_map, l2_map, sana_map, matriculas_map, reg_code_map = maps
    reg = str(row.get("Matricula", "")).strip().upper()
    flight_code = str(row.get("prefix3", "")).strip().upper()
    external_code = external_codes.get(reg)
    code, discrepancy = resolve_operator_code(reg, flight_code, external_code, reg_code_map)

    operator = matriculas_map.get(normalize_registration(reg), "") or icao_map.get(code, "")
    if code in l1_map:
        x = l1_map[code]
        return CrossResult("Layer 1", str(x["operator"] or operator), x["done"], x["objective"], x["remaining"], fmt_date(x["last"]), external_code or "", discrepancy, code)
    if code in sana_map:
        x = sana_map[code]
        return CrossResult("SANA", str(x["operator"] or operator), x["done"], x["objective"], x["remaining"], fmt_date(x["last"]), external_code or "", discrepancy, code)
    if code in l2_map:
        x = l2_map[code]
        return CrossResult("Layer 2", str(x["operator"] or operator), x["done"], x["objective"], x["remaining"], fmt_date(x["last"]), external_code or "", discrepancy, code)
    return CrossResult("No encontrado", operator, None, None, None, "", external_code or "", discrepancy, code)


def enrich_flights(df: pd.DataFrame, maps) -> pd.DataFrame:
    codes = build_external_operator_map(df["Matricula"].tolist())
    results = df.apply(lambda row: cross_reference(row, maps, codes), axis=1)
    out = df.copy()
    out["Código externo"] = [r.codigo_externo for r in results]
    out["Operador (maestro)"] = [r.operador for r in results]
    out["Tipo objetivo"] = [r.tipo for r in results]
    out["Inspecciones realizadas"] = [to_int_or_none(r.inspecciones) for r in results]
    out["Objetivo 2026"] = [to_int_or_none(r.objetivo) for r in results]
    out["Restantes"] = [to_int_or_none(r.restantes) for r in results]
    out["Última inspección"] = [r.ultima for r in results]
    out["_discrepancia"] = [r.discrepancia for r in results]
    out["_codigo_usado"] = [r.codigo_usado for r in results]
    return out[OUTPUT_COLUMNS + ["_discrepancia", "_codigo_usado"]]


def build_excel(df: pd.DataFrame, fecha: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cruce"
    headers = OUTPUT_COLUMNS
    last_col = get_column_letter(1 + len(headers))
    ws.merge_cells(f"B2:{last_col}2")
    ws["B2"] = f"Cruce tráfico NOP + Objetivos SAFA/SACA/SANA ({fecha})"
    ws["B2"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["B2"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["B2"].alignment = Alignment(horizontal="center")
    border = Border(*(Side(style="thin", color="D9D9D9") for _ in range(4)))
    for col, header in enumerate(headers, 2):
        cell = ws.cell(4, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    centered = {"Hora", "ARCID", "Aeronave", "Matricula", "ADEP", "ADES", "prefix3", "Código externo", "Tipo objetivo", "Inspecciones realizadas", "Objetivo 2026", "Restantes", "Última inspección"}
    orange_font = Font(size=10, bold=True, color="E36C0A")
    normal_font = Font(size=10)
    for r, (_, row) in enumerate(df.iterrows(), 5):
        for c, header in enumerate(headers, 2):
            value = to_int_or_none(row[header]) if header in INTEGER_COLUMNS else row[header]
            if value is None or (isinstance(value, float) and pd.isna(value)):
                value = None
            cell = ws.cell(r, c, value)
            cell.font = orange_font if bool(row.get("_discrepancia", False)) else normal_font
            cell.alignment = Alignment(horizontal="center" if header in centered else "left")
            cell.border = border
            if header in INTEGER_COLUMNS and value is not None:
                cell.number_format = "0"
    widths = {"Hora": 8, "ARCID": 12, "Aeronave": 11, "Matricula": 12, "ADEP": 8, "ADES": 8, "prefix3": 9, "Código externo": 14, "Operador (maestro)": 42, "Tipo objetivo": 14, "Inspecciones realizadas": 12, "Objetivo 2026": 12, "Restantes": 10, "Última inspección": 16}
    for c, h in enumerate(headers, 2):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(h, 14)
    ws.freeze_panes = "B5"
    ws.auto_filter.ref = f"B4:{last_col}{4 + len(df)}"
    note = 7 + len(df)
    ws.cell(note, 2, "Prioridad de operador: matrícula en Excel maestro > hexdb.io > código de vuelo.").font = Font(size=8, italic=True, color="808080")
    ws.cell(note + 1, 2, "Texto naranja: discrepancia entre el código seleccionado y el código del vuelo o hexdb.io.").font = Font(size=8, italic=True, color="808080")
    b = BytesIO(); wb.save(b); b.seek(0); return b


def build_pdf(df: pd.DataFrame, fecha: str) -> BytesIO:
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A3), leftMargin=8*mm, rightMargin=8*mm, topMargin=8*mm, bottomMargin=8*mm)
    styles = getSampleStyleSheet()
    normal = ParagraphStyle("n", parent=styles["Normal"], fontSize=7, leading=8)
    orange = ParagraphStyle("o", parent=normal, textColor=colors.HexColor("#E36C0A"))
    story = [Paragraph(f"Cruce tráfico NOP + Objetivos SAFA/SACA/SANA ({fecha})", styles["Heading1"]), Spacer(1, 4*mm)]
    for start in range(0, len(df), 25):
        chunk = df.iloc[start:start+25]
        table_data = [[Paragraph(h, normal) for h in OUTPUT_COLUMNS]]
        for _, row in chunk.iterrows():
            style = orange if bool(row.get("_discrepancia", False)) else normal
            vals = []
            for h in OUTPUT_COLUMNS:
                v = to_int_or_none(row[h]) if h in INTEGER_COLUMNS else row[h]
                vals.append(Paragraph("" if v is None or (isinstance(v, float) and pd.isna(v)) else str(v), style))
            table_data.append(vals)
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#4F81BD")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#D9D9D9")),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ]))
        story.append(table)
        if start + 25 < len(df): story.append(PageBreak())
    doc.build(story); buf.seek(0); return buf


def apply_filters(df: pd.DataFrame) -> pd.DataFrame:
    st.markdown("### Filtros antes de descargar")
    c1, c2, c3 = st.columns(3)
    with c1: search = st.text_input("ARCID (texto libre)", "")
    with c2: ades = st.multiselect("ADES (destino)", sorted(x for x in df["ADES"].dropna().unique() if x))
    with c3: ops = st.multiselect("Operador", sorted(x for x in df["Operador (maestro)"].dropna().unique() if x))
    c4, c5, c6 = st.columns(3)
    with c4: types = st.multiselect("Tipo objetivo", sorted(df["Tipo objetivo"].unique()), default=sorted(df["Tipo objetivo"].unique()))
    with c5:
        nums = pd.to_numeric(df["Restantes"], errors="coerce")
        mx = int(nums.max()) if nums.notna().any() else 1
        rest = st.slider("Restantes (rango)", 0, max(1, mx), (0, max(1, mx)))
    with c6:
        period = st.selectbox("Última inspección", ["Todas las fechas", "Última semana", "Último mes", "No en la última semana", "No en el último mes"])
    only_discrepancy = st.checkbox("Mostrar solo discrepancias (operador ≠ código de vuelo)")
    out = df.copy()
    if search: out = out[out["ARCID"].str.contains(search, case=False, na=False)]
    if ades: out = out[out["ADES"].isin(ades)]
    if ops: out = out[out["Operador (maestro)"].isin(ops)]
    if types: out = out[out["Tipo objetivo"].isin(types)]
    nums = pd.to_numeric(out["Restantes"], errors="coerce")
    out = out[nums.isna() | nums.between(rest[0], rest[1])]
    dates = pd.to_datetime(out["Última inspección"], errors="coerce")
    today = datetime.now().date()
    if period == "Última semana": out = out[dates.notna() & (dates.dt.date >= today - pd.Timedelta(days=7)) & (dates.dt.date <= today)]
    elif period == "Último mes": out = out[dates.notna() & (dates.dt.date >= today - pd.Timedelta(days=30)) & (dates.dt.date <= today)]
    elif period == "No en la última semana": out = out[dates.isna() | (dates.dt.date < today - pd.Timedelta(days=7))]
    elif period == "No en el último mes": out = out[dates.isna() | (dates.dt.date < today - pd.Timedelta(days=30))]
    if only_discrepancy: out = out[out["_discrepancia"]]
    return out


def render_app():
    st.title(APP_TITLE); st.caption(APP_CAPTION)
    c1, c2 = st.columns(2)
    with c1: pdf_file = st.file_uploader("1. PDF de tráfico (NOP / ARCID)", type=["pdf"])
    with c2: xlsx_file = st.file_uploader("2. Excel maestro", type=["xlsx"])
    if st.button("Generar cruce", type="primary", disabled=not(pdf_file and xlsx_file)):
        flights, undetected = parse_pdf_flights(pdf_file.read())
        result = enrich_flights(flights, build_master_maps(xlsx_file.read()))
        st.session_state["result"] = result
        st.session_state["undetected"] = undetected
        st.session_state["date"] = datetime.now().strftime("%Y%m%d")
    if "result" not in st.session_state:
        st.info("Sube ambos archivos y pulsa 'Generar cruce'."); return
    result = st.session_state["result"]
    undetected = st.session_state["undetected"]
    if not undetected.empty:
        with st.expander(f"Vuelos no detectados ({len(undetected)})"):
            st.dataframe(undetected.rename(columns={"raw":"Texto original", "motivo":"Motivo"}), use_container_width=True)
    filtered = apply_filters(result)
    display = filtered[OUTPUT_COLUMNS].copy()
    def style_row(row):
        return ["color:#E36C0A;font-weight:bold"] * len(row) if filtered.loc[row.name, "_discrepancia"] else [""] * len(row)
    st.dataframe(display.style.apply(style_row, axis=1), use_container_width=True, height=500)
    date = st.session_state["date"]
    a, b = st.columns(2)
    with a: st.download_button("Descargar Excel filtrado", build_excel(filtered, date), f"GCTS_{date}_filtrado.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with b: st.download_button("Descargar PDF filtrado", build_pdf(filtered, date), f"GCTS_{date}_filtrado.pdf", "application/pdf")

if __name__ == "__main__":
    render_app()
